"""Unit tests for the SpreadsheetBench OpenReward environment.

Tests cover:
- Evaluation logic (transform_value, compare_cell_value, cell range helpers)
- Workbook comparison with in-memory openpyxl workbooks
- Task structure and data integrity
- Environment class methods
"""

import datetime
import io
import json
from pathlib import Path

import openpyxl
import pytest

from evaluate import (
    col_name_to_num,
    col_num_to_name,
    compare_cell_value,
    compare_workbooks,
    datetime_to_float,
    generate_cell_names,
    parse_answer_position,
    transform_value,
)


# ---------------------------------------------------------------------------
# Value transformation tests
# ---------------------------------------------------------------------------

class TestTransformValue:
    def test_none(self):
        assert transform_value(None) is None

    def test_int(self):
        assert transform_value(42) == 42.0

    def test_float_rounding(self):
        assert transform_value(3.14159) == 3.14

    def test_float_round_half(self):
        assert transform_value(1.005) == 1.0  # banker's rounding

    def test_zero(self):
        assert transform_value(0) == 0.0

    def test_negative_float(self):
        assert transform_value(-2.567) == -2.57

    def test_bool_passthrough(self):
        # bool is a subclass of int, should be preserved
        assert transform_value(True) is True
        assert transform_value(False) is False

    def test_string_numeric(self):
        assert transform_value("3.14159") == 3.14

    def test_string_non_numeric(self):
        assert transform_value("hello") == "hello"

    def test_string_empty(self):
        assert transform_value("") == ""

    def test_datetime(self):
        dt = datetime.datetime(2024, 1, 15, 12, 0, 0)
        result = transform_value(dt)
        # Should be a float (Excel serial date), rounded to 0 decimals
        assert isinstance(result, float)
        assert result == round(datetime_to_float(dt), 0)

    def test_time(self):
        t = datetime.time(12, 30, 0)
        result = transform_value(t)
        # str(time(12,30,0)) = "12:30:00", [:-3] removes ":00" -> "12:30"
        assert result == "12:30"

    def test_time_with_seconds(self):
        t = datetime.time(12, 30, 45)
        result = transform_value(t)
        # str(time(12,30,45)) = "12:30:45", [:-3] removes ":45" -> "12:30"
        assert result == "12:30"


# ---------------------------------------------------------------------------
# Cell value comparison tests
# ---------------------------------------------------------------------------

class TestCompareCellValue:
    def test_equal_ints(self):
        assert compare_cell_value(42, 42) is True

    def test_equal_floats(self):
        assert compare_cell_value(3.14, 3.14) is True

    def test_equal_strings(self):
        assert compare_cell_value("hello", "hello") is True

    def test_none_none(self):
        assert compare_cell_value(None, None) is True

    def test_none_empty_string(self):
        assert compare_cell_value(None, "") is True

    def test_empty_string_none(self):
        assert compare_cell_value("", None) is True

    def test_empty_strings(self):
        assert compare_cell_value("", "") is True

    def test_different_values(self):
        assert compare_cell_value(1, 2) is False

    def test_different_strings(self):
        assert compare_cell_value("hello", "world") is False

    def test_type_mismatch_after_transform(self):
        # "abc" stays string, 1 becomes float — different types
        assert compare_cell_value("abc", 1) is False

    def test_string_number_vs_number(self):
        # "3.14" transforms to 3.14 (float), 3.14 transforms to 3.14 (float)
        assert compare_cell_value("3.14", 3.14) is True

    def test_rounding_makes_equal(self):
        # Both round to 1.23
        assert compare_cell_value(1.234, 1.231) is True

    def test_rounding_makes_different(self):
        assert compare_cell_value(1.234, 1.239) is False


# ---------------------------------------------------------------------------
# Column helpers tests
# ---------------------------------------------------------------------------

class TestColumnHelpers:
    def test_col_name_to_num(self):
        assert col_name_to_num("A") == 1
        assert col_name_to_num("Z") == 26
        assert col_name_to_num("AA") == 27
        assert col_name_to_num("AB") == 28
        assert col_name_to_num("AZ") == 52

    def test_col_num_to_name(self):
        assert col_num_to_name(1) == "A"
        assert col_num_to_name(26) == "Z"
        assert col_num_to_name(27) == "AA"
        assert col_num_to_name(28) == "AB"

    def test_roundtrip(self):
        for i in range(1, 100):
            assert col_name_to_num(col_num_to_name(i)) == i


# ---------------------------------------------------------------------------
# Cell range generation tests
# ---------------------------------------------------------------------------

class TestGenerateCellNames:
    def test_single_cell(self):
        assert generate_cell_names("A1") == ["A1"]

    def test_single_cell_lowercase(self):
        assert generate_cell_names("a1") == ["A1"]

    def test_range_single_column(self):
        assert generate_cell_names("A1:A3") == ["A1", "A2", "A3"]

    def test_range_single_row(self):
        assert generate_cell_names("A1:C1") == ["A1", "B1", "C1"]

    def test_range_columns_first(self):
        # Columns iterate first, then rows within each column
        result = generate_cell_names("A1:C3")
        expected = ["A1", "A2", "A3", "B1", "B2", "B3", "C1", "C2", "C3"]
        assert result == expected

    def test_range_2x2(self):
        result = generate_cell_names("A1:B2")
        expected = ["A1", "A2", "B1", "B2"]
        assert result == expected


# ---------------------------------------------------------------------------
# Answer position parsing tests
# ---------------------------------------------------------------------------

class TestParseAnswerPosition:
    def test_simple_range(self):
        result = parse_answer_position("A1:B5")
        assert result == [(None, "A1:B5")]

    def test_sheet_qualified(self):
        result = parse_answer_position("Sheet1!A1:B5")
        assert result == [("Sheet1", "A1:B5")]

    def test_quoted_sheet(self):
        result = parse_answer_position("'Sheet 1'!A1:B5")
        assert result == [("Sheet 1", "A1:B5")]

    def test_comma_separated(self):
        result = parse_answer_position("Sheet1!A1:B5,Sheet2!C3:D10")
        assert result == [("Sheet1", "A1:B5"), ("Sheet2", "C3:D10")]

    def test_mixed_qualified_and_not(self):
        result = parse_answer_position("Sheet1!A1:B5,C3:D10")
        assert result == [("Sheet1", "A1:B5"), (None, "C3:D10")]

    def test_single_cell_with_sheet(self):
        result = parse_answer_position("Sheet1!A1")
        assert result == [("Sheet1", "A1")]

    def test_fullwidth_colon(self):
        result = parse_answer_position("G12\uff1aJ15")
        assert result == [(None, "G12:J15")]

    def test_nbsp_stripped(self):
        result = parse_answer_position("\xa0'Sheet 2'!B2:B8")
        assert result == [("Sheet 2", "B2:B8")]

    def test_quoted_cell_range(self):
        # Some answer_positions have stray quotes around the cell range
        result = parse_answer_position("'CS!'B1:B8")
        assert result == [("CS", "B1:B8")]


# ---------------------------------------------------------------------------
# Workbook comparison tests (using in-memory workbooks)
# ---------------------------------------------------------------------------

def _create_workbook(data: dict[str, dict[str, object]]) -> bytes:
    """Create an xlsx workbook in-memory from {sheet_name: {cell_ref: value}} dict."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    for sheet_name, cells in data.items():
        ws = wb.create_sheet(sheet_name)
        for cell_ref, value in cells.items():
            ws[cell_ref] = value
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestCompareWorkbooks:
    def test_matching_workbooks(self, tmp_path):
        data = {"Sheet1": {"A1": 42, "A2": "hello", "B1": 3.14}}
        gt_bytes = _create_workbook(data)
        proc_bytes = _create_workbook(data)

        gt_path = tmp_path / "gt.xlsx"
        gt_path.write_bytes(gt_bytes)

        assert compare_workbooks(gt_path, proc_bytes, "A1:B2") is True

    def test_mismatched_values(self, tmp_path):
        gt_data = {"Sheet1": {"A1": 42, "A2": "hello"}}
        proc_data = {"Sheet1": {"A1": 99, "A2": "hello"}}

        gt_path = tmp_path / "gt.xlsx"
        gt_path.write_bytes(_create_workbook(gt_data))

        assert compare_workbooks(gt_path, _create_workbook(proc_data), "A1:A2") is False

    def test_missing_sheet_in_proc(self, tmp_path):
        gt_data = {"Sheet1": {"A1": 42}}
        proc_data = {"Sheet2": {"A1": 42}}

        gt_path = tmp_path / "gt.xlsx"
        gt_path.write_bytes(_create_workbook(gt_data))

        assert compare_workbooks(gt_path, _create_workbook(proc_data), "Sheet1!A1") is False

    def test_multi_sheet_comparison(self, tmp_path):
        data = {"Sheet1": {"A1": 1, "A2": 2}, "Sheet2": {"B1": 3, "B2": 4}}

        gt_path = tmp_path / "gt.xlsx"
        gt_path.write_bytes(_create_workbook(data))

        assert compare_workbooks(
            gt_path, _create_workbook(data), "Sheet1!A1:A2,Sheet2!B1:B2"
        ) is True

    def test_none_vs_empty_matches(self, tmp_path):
        gt_data = {"Sheet1": {"A1": None}}
        proc_data = {"Sheet1": {"A1": ""}}

        gt_path = tmp_path / "gt.xlsx"
        gt_path.write_bytes(_create_workbook(gt_data))

        assert compare_workbooks(gt_path, _create_workbook(proc_data), "A1") is True

    def test_float_rounding_matches(self, tmp_path):
        gt_data = {"Sheet1": {"A1": 1.234}}
        proc_data = {"Sheet1": {"A1": 1.231}}

        gt_path = tmp_path / "gt.xlsx"
        gt_path.write_bytes(_create_workbook(gt_data))

        # Both round to 1.23
        assert compare_workbooks(gt_path, _create_workbook(proc_data), "A1") is True

    def test_corrupt_proc_file(self, tmp_path):
        gt_data = {"Sheet1": {"A1": 42}}
        gt_path = tmp_path / "gt.xlsx"
        gt_path.write_bytes(_create_workbook(gt_data))

        assert compare_workbooks(gt_path, b"not a valid xlsx", "A1") is False


# ---------------------------------------------------------------------------
# Task structure tests (require prepare_data.py to have been run)
# ---------------------------------------------------------------------------

class TestTaskStructure:
    @pytest.fixture(autouse=True)
    def _load_data(self):
        data_path = Path(__file__).parent / "server_data" / "spreadsheetbench" / "dataset.json"
        if not data_path.exists():
            pytest.skip("dataset.json not found — run prepare_data.py first")
        with open(data_path) as f:
            self.dataset = json.load(f)

    def test_total_count(self):
        # 912 original tasks minus those excluded by validation (broken metadata)
        assert len(self.dataset) == 905

    def test_unique_ids(self):
        ids = [str(t["id"]) for t in self.dataset]
        assert len(ids) == len(set(ids))

    def test_required_fields(self):
        required = {"id", "instruction", "instruction_type", "answer_position", "num_test_cases"}
        for task in self.dataset:
            assert required.issubset(set(task.keys())), f"Task {task['id']} missing fields"

    def test_valid_instruction_types(self):
        valid = {"Cell-Level Manipulation", "Sheet-Level Manipulation"}
        for task in self.dataset:
            assert task["instruction_type"] in valid, f"Task {task['id']} has invalid type"

    def test_test_case_counts(self):
        for task in self.dataset:
            assert task["num_test_cases"] >= 1, f"Task {task['id']} has no test cases"

    def test_total_test_cases(self):
        total = sum(t["num_test_cases"] for t in self.dataset)
        # Original benchmark has ~2729 test cases, minus excluded tasks
        assert total > 2600, f"Expected ~2700 total test cases, got {total}"

    def test_stable_ordering(self):
        ids = [str(t["id"]) for t in self.dataset]
        assert ids == sorted(ids)


# ---------------------------------------------------------------------------
# Environment class tests
# ---------------------------------------------------------------------------

class TestSpreadsheetBenchEnv:
    @pytest.fixture(autouse=True)
    def _check_data(self):
        data_path = Path(__file__).parent / "server_data" / "spreadsheetbench" / "dataset.json"
        if not data_path.exists():
            pytest.skip("dataset.json not found — run prepare_data.py first")

    def test_list_splits(self):
        from spreadsheetbench import SpreadsheetBench
        assert SpreadsheetBench.list_splits() == ["test"]

    def test_list_tasks_test(self):
        from spreadsheetbench import SpreadsheetBench
        tasks = SpreadsheetBench.list_tasks("test")
        assert len(tasks) == 905
        for task in tasks[:10]:
            assert "id" in task
            assert "instruction_type" in task
            # Should NOT expose server-side fields
            assert "instruction" not in task
            assert "answer_position" not in task

    def test_list_tasks_unknown_split(self):
        from spreadsheetbench import SpreadsheetBench
        assert SpreadsheetBench.list_tasks("train") == []

    def test_list_tasks_stable_ordering(self):
        from spreadsheetbench import SpreadsheetBench
        tasks1 = SpreadsheetBench.list_tasks("test")
        tasks2 = SpreadsheetBench.list_tasks("test")
        assert [t["id"] for t in tasks1] == [t["id"] for t in tasks2]
