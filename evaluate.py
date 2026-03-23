"""SpreadsheetBench evaluation — cell-level comparison of spreadsheet workbooks.

Ported from the original SpreadsheetBench evaluation code:
https://github.com/RUCKBReasoning/SpreadsheetBench/blob/main/evaluation/evaluation.py

Compares cell values at specified positions between a ground-truth workbook and
a processed (agent-produced) workbook. Only cell values are compared — colors
and formatting are ignored.
"""

import datetime
import io
import re
from pathlib import Path

import openpyxl


def datetime_to_float(dt: datetime.datetime) -> float:
    """Convert a datetime to Excel serial date (days since 1899-12-30)."""
    epoch = datetime.datetime(1899, 12, 30)
    delta = dt - epoch
    return delta.days + delta.seconds / 86400.0


def transform_value(v):
    """Normalize a cell value for comparison.

    - int/float: round to 2 decimal places
    - datetime.time: convert to string, truncate last 3 chars
    - datetime.datetime: convert to Excel serial date, round to 0 decimals
    - str: try float conversion (round to 2dp), otherwise keep as-is
    - None: pass through
    """
    if v is None:
        return v
    if isinstance(v, bool):
        # bool is a subclass of int, handle before int check
        return v
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if isinstance(v, datetime.time):
        return str(v)[:-3]
    if isinstance(v, datetime.datetime):
        return round(datetime_to_float(v), 0)
    if isinstance(v, str):
        try:
            return round(float(v), 2)
        except ValueError:
            return v
    return v


def compare_cell_value(v1, v2) -> bool:
    """Compare two cell values after transformation.

    None and empty string are treated as equivalent.
    After that, strict type matching is applied followed by exact equality.
    """
    v1 = transform_value(v1)
    v2 = transform_value(v2)

    # None / empty-string equivalence
    if v1 == "" and v2 is None:
        return True
    if v1 is None and v2 == "":
        return True
    if v1 is None and v2 is None:
        return True
    if v1 == "" and v2 == "":
        return True

    # Strict type check
    if type(v1) != type(v2):
        return False

    return v1 == v2


# ---------------------------------------------------------------------------
# Cell-range helpers
# ---------------------------------------------------------------------------

def col_name_to_num(name: str) -> int:
    """Convert a column name like 'A', 'Z', 'AA', 'AB' to a 1-based number."""
    num = 0
    for ch in name.upper():
        num = num * 26 + (ord(ch) - ord("A") + 1)
    return num


def col_num_to_name(n: int) -> str:
    """Convert a 1-based column number to a column name like 'A', 'AB'."""
    result = []
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result.append(chr(65 + remainder))
    return "".join(reversed(result))


def parse_cell_ref(cell_ref: str) -> tuple[str, int]:
    """Parse 'AB12' into ('AB', 12)."""
    match = re.match(r"^([A-Za-z]+)(\d+)$", cell_ref)
    if not match:
        raise ValueError(f"Invalid cell reference: {cell_ref}")
    return match.group(1).upper(), int(match.group(2))


def generate_cell_names(range_str: str) -> list[str]:
    """Expand a cell range like 'A1:C3' into individual cell names.

    Iteration order: columns first, then rows within each column.
    E.g. 'A1:C3' -> ['A1','A2','A3','B1','B2','B3','C1','C2','C3']

    Single cell (no colon): 'A1' -> ['A1']
    """
    range_str = range_str.strip()
    if ":" not in range_str:
        return [range_str.upper()]

    start, end = range_str.split(":")
    start_col, start_row = parse_cell_ref(start)
    end_col, end_row = parse_cell_ref(end)

    start_col_num = col_name_to_num(start_col)
    end_col_num = col_name_to_num(end_col)

    cells = []
    for col_num in range(start_col_num, end_col_num + 1):
        col_name = col_num_to_name(col_num)
        for row in range(start_row, end_row + 1):
            cells.append(f"{col_name}{row}")
    return cells


def parse_answer_position(answer_position: str) -> list[tuple[str | None, str]]:
    """Parse an answer_position string into (sheet_name, cell_range) pairs.

    Handles:
    - Comma-separated ranges: "Sheet1!A1:B5,Sheet2!C3:D10"
    - Sheet-qualified: "Sheet1!A1:B5" -> ("Sheet1", "A1:B5")
    - Unqualified: "A1:B5" -> (None, "A1:B5") — caller uses default sheet
    - Quoted sheet names: "'Sheet 1'!A1:B5" -> ("Sheet 1", "A1:B5")
    - Full-width colons: "G12：J15" -> normalized to "G12:J15"
    - Non-breaking spaces: stripped from input
    """
    # Normalize: replace full-width colon with regular colon, strip NBSP
    answer_position = answer_position.replace("\uff1a", ":").replace("\xa0", " ").strip()

    results = []
    for part in answer_position.split(","):
        part = part.strip()
        if not part:
            continue
        if "!" in part:
            sheet_name, cell_range = part.split("!", 1)
            # Strip surrounding quotes from sheet name and cell range
            # Handles formats like: 'Sheet Name'!A1:B5  and  'CS!'B1:B8
            sheet_name = sheet_name.strip("'\"")
            cell_range = cell_range.strip("'\"")
            results.append((sheet_name, cell_range))
        else:
            # Strip any surrounding quotes from unqualified ranges too
            results.append((None, part.strip("'\"")))
    return results


def _resolve_sheet_name(wb: openpyxl.Workbook, name: str) -> str | None:
    """Resolve a sheet name case-insensitively. Returns the actual name or None."""
    # Exact match first
    if name in wb.sheetnames:
        return name
    # Case-insensitive fallback
    name_lower = name.lower()
    for sn in wb.sheetnames:
        if sn.lower() == name_lower:
            return sn
    return None


def cell_level_compare(
    wb_gt: openpyxl.Workbook,
    wb_proc: openpyxl.Workbook,
    sheet_name: str,
    cell_range: str,
) -> bool:
    """Compare cell values in a given range between two workbooks."""
    gt_sheet = _resolve_sheet_name(wb_gt, sheet_name)
    proc_sheet = _resolve_sheet_name(wb_proc, sheet_name)
    if gt_sheet is None or proc_sheet is None:
        return False

    ws_gt = wb_gt[gt_sheet]
    ws_proc = wb_proc[proc_sheet]

    cells = generate_cell_names(cell_range)
    for cell_ref in cells:
        gt_val = ws_gt[cell_ref].value
        proc_val = ws_proc[cell_ref].value
        if not compare_cell_value(gt_val, proc_val):
            return False
    return True


def compare_workbooks(
    gt_path: str | Path,
    proc_data: bytes | str | Path,
    answer_position: str,
) -> bool:
    """Compare a ground-truth workbook against a processed workbook.

    Args:
        gt_path: Path to the ground-truth (answer) xlsx file.
        proc_data: Either bytes (downloaded from sandbox), or a path to the
            processed (agent-produced) xlsx file.
        answer_position: Cell position spec, e.g. "Sheet1!A1:B5,Sheet2!C3".

    Returns:
        True if all specified cells match, False otherwise.
    """
    try:
        wb_gt = openpyxl.load_workbook(str(gt_path), data_only=True)
    except Exception:
        return False

    try:
        if isinstance(proc_data, bytes):
            wb_proc = openpyxl.load_workbook(io.BytesIO(proc_data), data_only=True)
        else:
            wb_proc = openpyxl.load_workbook(str(proc_data), data_only=True)
    except Exception:
        return False

    parsed = parse_answer_position(answer_position)

    for sheet_name, cell_range in parsed:
        # If no sheet specified, use the first sheet from ground truth
        if sheet_name is None:
            sheet_name = wb_gt.sheetnames[0]

        if not cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range):
            return False

    return True
